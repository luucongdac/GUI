#!/bin/python
import codecs
import openpyxl
import sys, os
from time import gmtime, strftime
import datetime
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string
import array as arr
import numpy as np
import sys
import time
import os, shutil
from shutil import copyfile

#thread fun :)))
import threading

def clock():
    while True:  
        #print(datetime.datetime.now().strftime("%H:%M:%S"), end="\r")
        a = int(datetime.datetime.now().strftime("%S"))
        if (a % 2) == 0:
            print('----\\\\\\\----', end="\r")
        else:
            print('----////----', end="\r")
        time.sleep(0.1)
try:
    x = threading.Thread(target=clock, daemon = True)
    x.start()
except:
    print('Can not create Thread')

#-----------------------------------------------------------------
now = datetime.datetime.now()
mins = now.strftime("%M")
sec = now.strftime("%S")

print('\n' +'\n' + str(now) + '\n')

if len(sys.argv) >1:
    t_t = str(sys.argv[1])
    t_t = t_t.replace('\\','/')
    t_t = t_t.replace('\xa0','')
    t_t = t_t.replace('\u202a','')
    t_t = t_t.replace(' ','')
    t_t = t_t.replace('"','')
    t_t = t_t.replace('\'','')
    
    t_t_t = str(sys.argv[2])
    t_t_t = t_t_t.replace(' ','')
    print('\n')
    print('Path Report: ' + t_t)
    print('\n')
    print('Source function: ' + t_t_t)

#t_t = "D:/SVN HilCS/Deliverables/Branches/Task00129_P33A_QCV2/Task00xxx_GroupX_osr_pre.c_DacLuu/単体テスト仕様書/vOSR_Pre_ObjectDataUpdate.xlsx"
#t_t_t = "asdasd"

#-copy temp and wriite end line
def add_last_character(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb['テストケース表']
    max_rown = sheet.max_row
    sheet['B' + str(max_rown +1)] = 'end'
    wb.save(path)

tmp1 = t_t.split('/')
des_excel_copy = 'D:/' + tmp1[-1]
try:
    shutil.copy(t_t, 'D:/')
except:
    os.remove(des_excel_copy)
try:
    shutil.copy(t_t, 'D:/')
except:
    print('ReCopy Fail')
add_last_character(des_excel_copy)



#

log_flag = 0
#print('Print log in console ? (0 or 1)')
#log_flag = int(input())
# create file log
try:
    try:
        file_test_log = open('D:/Testlog.txt','r+')
        file_test_log.truncate(0)
    except:
        file_test_log = open('F:/Testlog.txt','r+')
        file_test_log.truncate(0)
except:
    try:
        file_test_log = open('D:/Testlog.txt','w+')
    except:
        file_test_log = open('F:/Testlog.txt','w+')


def log(a):
    a = a.replace('\xa0',' ')
    if log_flag == 1:
        print(a)
    log_t = a + '\n'
    file_test_log.write(log_t)

log('\n' + str(now) + '\n')
log('-----------------------------------------------------')
#-------------------------------------
#print('Please type 1 if source in excel file and 0 to type')
#source_excel = int(input())
#source_excel = 0

new_old = 1
#-------------------------------------

#print('\n')

#excel = "C:/Users/daccl.hy/Desktop/1.xlsx"
excel = "D:/temp.xlsx"
source_name = ''
function_name = ''
duong_link = ''
function_csv = ''
source_excel = 0

if source_excel == 1:
    wb_t = openpyxl.load_workbook(excel)
    sheet_t = wb_t['Sheet1']
    excel = sheet_t["A7"].value
    excel = excel.replace("\\", '/')
    function_name = sheet_t["A13"].value + '/'
else:
    #print('Please add path testSpec')
    #t_t = input()
    excel = t_t.replace("\\", '/')
    excel = des_excel_copy
    print('\n')
    log('\n')
    #print('Please add name of Source.c test')
    #t_t_t = input()
    function_name = t_t_t + '/'
    print('\n')

#print('\n')
log(excel + '\n')
log(function_name + '\n')


log('-----------------------------------------------------')
wb = openpyxl.load_workbook(excel)
log('try to collect LOC of IF condition')
try:
    sheet_log = wb['カバレッジ結果']
    def colect_LOC_condition():
        a = 3
        list_colect_LOC = ''
        while True:
            a += 1
            st = str(sheet_log.cell(a,1).value) 
            if 'if(' in st:
                while True:
                    st = str(sheet_log.cell(a,1).value)
                    list_colect_LOC = list_colect_LOC + '$$$' + st
                    if ('{') in st:
                        break
                    a += 1
            if 'OMF Converter' in st:
                break
            if a > 2000:
                print('Fail sheet test log')
                break
        return list_colect_LOC
    list_colect_LOC = colect_LOC_condition()
    string_colect_LOC = list_colect_LOC.split('$$$')

    start = '('
    end = ')'
    for i in string_colect_LOC:
        if '(' in i:
            if ')' in i:
                if not 'MC/DC' in i:
                    trash = (i.split(start))[1].split(end)[0]
                    if not trash == '':
                        log( 'LOC' + '\t' + i[0:10] + '\t' + trash )
except:
    log('Fail to count LOC of condition')
log('End count LOC' + '\n')
log('-----------------------------------------------------')
#--------------------------------------------------------
file = "C://Users//daccl.hy//Desktop//auto//func_additional_signal_cal_IO.html"
#sheet.max_row
# quet row 3 tim ID va Commnent


sheet = wb['テストケース表']
#print('total column = ', sheet.max_column)
log('total column = ' + str(sheet.max_column))
#print('total row = ' ,sheet.max_row)
log('total row = ' + str(sheet.max_row))
#print('\n\n')
log('\n')
log('\n')
print('-------STARTING FIND TEST CASE-------')
log('-------STARTING FIND TEST CASE-------')
log('\n\n')

x = 1
while x <= sheet.max_column:
    if str(sheet.cell(3, x).value) == 'ID':
        break
    x += 1
ID = x
Comment = ID +1

# tim vi tri test case dau tien
x = 4
while x <= sheet.max_row:
    if not str(sheet.cell(x, 2).value) == 'None':
        break
    x += 1
##print(x)   vi tri bat dau cua testcase dau tien

# lay test case                                            
all_table = ['']*(sheet.max_row + 1)
while x <= sheet.max_row:
    if not str(sheet.cell(x, 1).value) == 'None':
        all_table[x] = str(sheet.cell(x, 1).value).replace('-','') + '$$$' + str(sheet.cell(x, 2).value) + '$$$' + str(sheet.cell(x, ID).value) + '$$$' + str(sheet.cell(x, Comment).value) + '$$$'
        t_t = all_table[x]
        all_table[x] = t_t.replace('\xa0','')
    x +=1
                                                                                                                                    
x =5
y = 0
while x <= sheet.max_row:
    if not str(sheet.cell(x, 2).value) == 'None':  
        all_table[x - y -1] = all_table[x - y -1] + str(sheet.cell(x -1, 1).value).replace('-','')
        t_t = all_table[x - y -1]
        all_table[x - y -1] = t_t.replace('\xa0','')
        y = 0
    else:
        y +=1
    x +=1

#--------------xoa None----------------------

tring_t = ''
tring_g =''
result_t = ['']*(sheet.max_row + 1)
x = 4
y = 0
while x <= sheet.max_row:
    tring_g = all_table[x].split('$$$')
    ##print(tring_g)
    if len(all_table[x]) >4:
        if not tring_g[1] == 'None':
            result_t[y] = all_table[x]
            y +=1  
    x +=1

#x =0
#while x <= sheet.max_row:
    ##print(result_t[x])
#    x +=1
#result_t[y] = tring_g[0] + ',' + tring_g[3] + ',' + tring_g[1] + ',' + tring_g[2]
x = 0
i = 0
array_testcase = ''  # all name test case
while x <= sheet.max_row:
    tring_g = result_t[x].split('$$$')
    ##print(len(tring_g))
    if len(tring_g) > 2:
        result_t[x] = tring_g[0] + '$' + tring_g[4] + '$' + tring_g[1] + '$' + tring_g[2]+ '$'+ tring_g[3]
        array_testcase =  array_testcase + ',' + tring_g[1]
        i +=1
    x +=1
# all var test case
#x =0
line_of_test = i
#while x <= sheet.max_row:
    ##print(result_t[x])
#    x +=1
#-------------------

max_row = sheet.max_row
line_of_test = sheet.max_row 
max_column = sheet.max_column
test_Anal = ''

string_result = result_t.copy()
string_copy = result_t.copy()
string_result_1  = string_copy.copy()
p = 0
while p < (y + 1):
    ##print(p, '\t' ,string_result_1[p].replace('$', '\t'))
        
    log(str(p) + '\t' + string_result_1[p].replace('$', '\t'))
    p += 1
#print('\n\n')
#print('\n\n')
log('\n')
log('\n')

# delete None
p = 0

#print('\n\n')
#print('\n\n')
log('\n\n')
#print('------------All test point following 14 point ---------')
log('--------All test point following 14 point ---------')
#print('\n')
log('\n')

all_test_poin_1 = ''
if new_old == 0:
    point_check = ['coverage_p1', 'div_zero_p2', 'overflow_p3', 'casting_p4', 'underflow_p5','array_p6', 'pointer_p7']
else:
    point_check = ['input_variable_p1', 'input_func_return_p2', 'condition_p3', 'sw_case_p4', 'zero_division_p5','calc_overflow_p6', 'casting_overflow_p8', 'array_p9', 'pointer_p10', 'loop_p11' ]
    
for k in point_check:
    ##print('-----------------------------', k, '-------------------------------')
    log('-----------------------------'+ str(k) + '--------------------------')
    p = 0
    h = 1
    string_t3 = ''
    while p < i:
        if  k in string_result[p]:
            string_result[p] = string_result[p]#.replace('-', '')
            string_t3 = string_result[p].split("$")
            ##print(string_t3)
            ##print( string_t3[2] + '\t' + string_t3[4] + '\t' + string_t3[0] + '~' + string_t3[1])
            t_t = string_t3[2] + '\t' + '\t' + '\t' + string_t3[4] + '\t' + string_t3[0] + '~' + string_t3[1]
            if k == 'input_variable_p1':
                all_test_poin_1 = all_test_poin_1 + '$' + string_t3[2]
            log(t_t)
            h += 1
        p +=1
    ##print('-------------------total = ',h-1)
    log('-------------------> total = '+ str(h-1))
    ##print('\n\n')
    log('\n\n')



#exit()
#wb.save(excel)
##print('\n')
log('\n')
##print('----------All input variable-------')
log('----------All input variable-------')
##print('\n')
log('\n')


sheet = wb['入出力データ分析表']

#-------------------------------------------------------------------------------------------------------
##print(sheet.max_column, sheet.max_row)
# find number of input
a = 4
while(a <= sheet.max_column):
    if str(sheet.cell(3, a).value) == 'None':
        a += 1
    else:
        break
number_colum_input = a
##print(number_colum_input)
# find Type of variable and colect
a = 1
while(a <= sheet.max_row):
    if str(sheet.cell(a, 1).value) == 'Type':
        cel_type = a
        break
    else:
        a += 1
##print(cel_type)        

type_var = ['']*(number_colum_input + 5)
type_var_output = ['']*(sheet.max_column + 1)

a = 3
while a <= (number_colum_input - 1):
    type_var[a] = str(sheet.cell(cel_type, a ).value)
    ##print(type_var[a])
    a +=1
# lay type ouput
a = number_colum_input
while a <= sheet.max_column:
    type_var_output[a] = str(sheet.cell(cel_type, a ).value)
    ##print(typr_var_output[a])
    a +=1
   
# dem so row cau variable va lay ten bien
a = 7
while(a <= sheet.max_row):
    if str(sheet.cell(a, 1).value) == 'None':
        a += 1
    else:
        break
number_row_var = a
##print(number_row_var)
var_name = ['']*(number_colum_input + 5)
a = 3
while a <= (number_colum_input - 1):
    b = 6
    while b <= (number_row_var -1):
        var_name[a] = var_name[a] + str(sheet.cell(b, a ).value)
        var_name[a] = var_name[a].replace('None','')
        var_name[a] = var_name[a].replace('AMSTB_SrcFile.c/','')
        var_name[a] = var_name[a].replace(function_name,'')
        # xoa them source c.c  'Sensor_Processing_sfunc.c/'
        b += 1
    a +=1
# lay ten bien ouput
var_name_output = ['']*(sheet.max_column + 1)
a = number_colum_input
while a <= sheet.max_column:
    b = 6
    while b <= (number_row_var -1):
        var_name_output[a] = var_name_output[a] + str(sheet.cell(b, a ).value)
        var_name_output[a] = var_name_output[a].replace('None','')
        var_name_output[a] = var_name_output[a].replace('AMSTB_SrcFile.c/','')
        var_name_output[a] = var_name_output[a].replace(function_name,'')
        # xoa them source c.c  'Sensor_Processing_sfunc.c/'
        b += 1
    a +=1

#----------
def delete_t(a):
    a = a.replace(' ','')
    a = a.replace('\t','')
    a = a.replace('\n','')
    a = a.replace('\r','')
    return a
# ghep type with name var  
a = 3
while a <= (number_colum_input - 1):
    type_var[a] = delete_t(type_var[a]) + ' '+ delete_t(var_name[a])
    #print(type_var[a])
    log(str(type_var[a]))
    a +=1
#print('\n')
log('\n')
#print('-----total = ', a-3,'--------------')
log('-----> total = '+ str(a-3) )
#print('\n')
log('\n\n')
#print('----------All ouput variable-------')
log('----------All ouput variable-------')
#print('\n')
#ghep type with namevar output
a = number_colum_input
while a <= sheet.max_column:
    type_var_output[a] = delete_t(type_var_output[a]) + ' '+ delete_t(var_name_output[a])
    #print(type_var_output[a])
    log(str(type_var_output[a]))
    a +=1
#print('\n')
#print('\n')
log('\n\n')
#print('-----total = ', a-number_colum_input,'--------------')
log('-----> total = '+ str(a-number_colum_input))
#print('\n')
log('\n')
#---------------------------
#print('-----All AMOUT in output var--------')
log('-----All AMOUT in output var--------')
#print('\n')
a = 0
for i in type_var_output:
    if '@AMOUT' in i:
        #print(i)
        log(str(i))
        a += 1
#print('-----------------------------------')
#print('\n')
log('\n')
#print('-----total = ', a,'--------------')
log('-----> total = '+ str(a))
#print('\n')
#print('\n')
log('\n')
# tim first Test Analysis
a = 1
while(a <= sheet.max_row):
    if not str(sheet.cell(a, 2).value) == 'Test\xa0Analysis\xa0Item':
        a += 1
    else:
        break
#---------------------------------------------------------------------------------------------------
first_colum = 3
last_colum = number_colum_input

first_row = a
last_row = sheet.max_row


#------------------------------------------------------------------------------------------------------

sheet = wb['入出力データ分析表']

#sheet.cell(row, column).value
x = first_colum
y = first_row

z = ['']*(last_colum + 1)

temp_t = ['']*(last_colum + 1)
while x <= last_colum:
    while y <= last_row:
        z[x] = z[x] + ',' +  str(sheet.cell(y, x).value)
        ##print(temp_t[x])
        y += 3
    y = first_row
    x += 1


array_t = ['']*(last_colum + 3)

a = 3
while a <= (last_colum):
    temp_t[a] = list(dict.fromkeys(z[a].split(',')))
    #temp_t[a] = temp_t[a].replace('None','')
    #my_array[a] = np.asarray(temp_t[a])
    #array_t[a].append(temp_t[a])
    ##print(temp_t[a])
    a +=1

my_array = ['']*(last_colum + 3)

a = 3
while a <= (last_colum):
    for i in temp_t[a]:      
       #my_array[a] = my_array[a].replace(',','')
       my_array[a] = my_array[a] + ',' + i
    a +=1
#print('--------all test var input-----------')
log('--------all test var input-----------')
#print('\n')
log('\n')

# sort test case following max -> min
a_raw = array_testcase
c_raw = a_raw.split(',')
def sort_name(string):
    b_raw = string
    tempt = ''
    if len(b_raw) > 0:
        d = b_raw.split(',')       
        for i in c_raw:
            if i in d:
               tempt = tempt + ',' + i
    return(tempt[1:])
#-----------------------------------------------------------------------------------------

a = 1
while a <= (last_colum):
    #my_array[a] = my_array[a].replace(',','')
    my_array[a] = my_array[a].replace(',,','')
    my_array[a] = my_array[a].replace(',None','')
    my_array[a] = my_array[a].replace('None','')
    my_array[a]= my_array[a]#.replace('-','')
    #my_array[a]= my_array[a].replace('#','')  ''''''''''''''''''''''''''
    #my_array[a] = str(a) + '\t' + my_array[a]
    my_array[a] =  my_array[a].replace(' ','')
    my_array[a] =  my_array[a].replace('\t','')
    my_array[a] =  my_array[a].replace('\r','')
    my_array[a] =  my_array[a].replace('\n','')
    a +=1

copy_tem_t = my_array.copy()
a = 1
while a <= (last_colum):
    my_array[a] = sort_name(copy_tem_t[a])
    my_array[a] = my_array[a].replace(',,','')
    my_array[a] = my_array[a].replace('-','')
    if a >= 3:
        #print(a-2, '\t',my_array[a])
        log(str(a-2) + '\t' + str(my_array[a]))
    a +=1

#---------------------------------------------------------
# colect name and test-case
p = 0

string_copy_name = ['']*(line_of_test + 1)
string_copy_testcase = ['']*(line_of_test + 1)
string_temp = ''
 
while p < (line_of_test + 1):
    string_temp = string_result_1[p].split("$")
    ##print(string_temp)
    if len(string_temp) >2:
        g = string_temp[2]
#    if not (g == 'A' and  g == 'B' and  g == 'C' and  g == 'D' and  g == 'E' and  g == 'F' and  g == 'G' and  g == 'H' and  g == 'I' and  g == 'J' 
#	and  g == 'K' and  g == 'L' and  g == 'M' and  g == 'N' and  g == 'O' and  g == 'P' and  g == 'Q' and  g == 'R' and  g == 'S' and  g == 'T' 
#	and  g == 'U' and  g == 'V' and  g == 'W' and  g == 'X' and  g == 'Y' and  g == 'Z' ):
        string_copy_name[p] = string_temp[2].replace('-','')
        #string_copy_name[p] = string_copy_name[p].replace('#','')''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
        string_copy_testcase[p] =string_temp[0].replace('-','') + '~' + string_temp[1].replace('-','')
        #string_copy_testcase[p] =string_temp[0] + '~' + string_temp[1]
        #string_copy_testcase[p] = string_copy_testcase[p].replace('#','')
    p +=1
#print('string_copy_name:     ', string_copy_name)
#print('string_copy_testcase: ', string_copy_testcase)
try: 
    log('string_copy_name:     ' + str(string_copy_name))
    log('string_copy_testcase:' + str(string_copy_testcase))
except:
    print('')
    
# danh lai testcase

my_array_t = ['']*last_colum
my_array_point1 = my_array.copy()
# delete var not labeled point 1  JUST for point 1 var
case_point_1 = all_test_poin_1.split('$')
i = 0
while i < len(case_point_1):
    case_point_1[i] = case_point_1[i].replace('-','')
    i += 1
#print(str(case_point_1))
#
#print(str(my_array_point1) + '\n')
y = 10
while y > 0:
    a = 1
    while a  <= (last_colum):
        i= my_array_point1[a]
        b = 0
        while b <= (line_of_test -1 ):
            p = string_copy_name[b]
            if (p in i):
                if len(p) >= y:
                    if p in case_point_1:
                        my_array_point1[a] = my_array_point1[a].replace(p,string_copy_testcase[b])   
                    else:
                        my_array_point1[a] = my_array_point1[a].replace(p,'777')      
            b += 1
        my_array_point1[a] = my_array_point1[a].replace('777,','')
        my_array_point1[a] = my_array_point1[a].replace(',777','')
        my_array_point1[a] = my_array_point1[a].replace('777','')
        a +=1
    y = y -1

#print(str(string_copy_name) + '\n')
#print(str(case_point_1) + '\n')

#print(str(my_array) + '\n')
# For all others var
y = 10
while y > 0:
    a = 1
    while a  <= (last_colum):
        i= my_array[a]
        b = 0
        while b <= (line_of_test -1 ):
            p = string_copy_name[b]
            if (p in i):
                if len(p) >= y:
                    my_array[a] = my_array[a].replace(p,string_copy_testcase[b])    
            b += 1
        a +=1
    y = y -1
#print(str(my_array) + '\n')
#
#-----------merge test case-----------
#a = '1~5,59~63,64~65,66~70,71~75,76~78,79~83,84~86,87~88,89~90,91~91,92~94,95~96,97~101, 222~256, 257~267, 277~277, 299~300'
#b =a.split(',')
def merger_modify(a):
    if len(a) == 0:
        return('')
    temp2 = ''
    temp3 = ''
    temp4 = ''
    b = a.split(',')
    temp1 = 0
    while temp1 < (len(b)-1):
        temp2 = b[temp1].split('~')
        temp3 = b[temp1+1].split('~')

        if (int(temp3[0]) - int(temp2[1])) == 1:
            b[temp1] = temp2[0] + '~0'
            b[temp1+1] = '0~' + temp3[1]
        temp1 += 1

    result = ''
    for i in b:
        result = result + ',' + i
    result = result.replace('0,0~','')

    return(result)


#-------------------------------------
# merge test case as  155~155

def merge_testcase_special(string):
    string = string.replace(' ','')
    if len(string) == 0:
        return('')
    a_raw = string.split(',')
    tempt = ''
    tempt_t = ''

    for i in a_raw:
        tempt = i.split('~')
        ##print(tempt)
        if len(tempt) == 2:
            if tempt[0] == tempt[1]:
                tempt_t = tempt_t + ',' +tempt[0]
            else:
                tempt_t = tempt_t + ',' +tempt[0] + '~' + tempt[1]
    return(tempt_t[1:])


#--------------------------
tt = ''
array_result_t = my_array.copy()
a = 1
while a <= (last_colum):
    tt = my_array[a]
    if len(tt) > 1:
        tt_t  = merger_modify(tt)
        tt_t = tt_t[1:]
        my_array[a] = merge_testcase_special(tt_t)
    ##print(my_array[a])
    a +=1
#print('\n\n')
log('\n')
# for case point 1
tt = ''
a = 1
while a <= (last_colum):
    tt = my_array_point1[a]
    if len(tt) > 1:
        tt_t  = merger_modify(tt)
        tt_t = tt_t[1:]
        my_array_point1[a] = merge_testcase_special(tt_t)
    ##print(my_array[a])
    a +=1
#print('\n\n')
log('\n')


#


type_var_sp = type_var.copy()
# ghep bien voi test case
a = 3
b = 3
while a <= (number_colum_input - 1):
    type_var[a] = type_var[a] + '\t' + my_array[b].replace(',', ', ')
    #my_array_point1
    type_var_sp[a] = type_var_sp[a] + '\t' + my_array_point1[b].replace(',', ', ')
    ##print(type_var[a])
    a +=1
    b +=1
#print(type_var)
#print(type_var_sp)
#print('\n\n')
log('\n')
log('\n')
#print('--------total input var =', a-3, '----------')
log('--------> total input var = '+ str(a-3))
#print('\n\n')
log('\n')
log('\n')
##print poin 1-2 new templace

point_check = ['Number', 'AMIN_return', 'AMOUT']
for k in point_check:
    #print('-----------------------------', k, '-------------------------------')
    log('-----------------------------'+ str(k) +'-------------------------------')
    a = 3
    b = 0
    while a <= (number_colum_input - 1):
        if k in type_var[a]:
            #print(type_var[a])
            log(str(type_var[a]))
            b +=1
        a +=1
    
    #print('-------------------total =--------',b , '--------------------------' )
    log('-------------------> total = '+ str(b))
    if k == 'AMOUT':
        log("WARNING- list all AMOUT var in INPUT to Point_2 ")
        log("Scroll up, list all AMOUT var in OUTPUT to Point_10 ")
    #print('\n\n')
    log('\n')
    log('\n')
    

#print('-----------------------------input variable-------------------------------')
log('-----------------------------all input variable is used-------------------------------')
a = 3
b = 0
while a <= (number_colum_input - 1):
    if (not('Number' in type_var[a]) and (not 'AMIN_return' in type_var[a])):
        #print(type_var[a])
        log(str(type_var[a]))
        b +=1
    a +=1
    
#print('-------------------total =--------',b , '--------------------------' )
log('-------------------> total = '+ str(b))
#print('\n\n')
log('\n')

log('-----------------------------input variable following label point 1-------------------------------')
a = 3
b = 0
while a <= (number_colum_input - 1):
    if (not('Number' in type_var_sp[a]) and (not 'AMIN_return' in type_var_sp[a])):
        #print(type_var[a])
        log(str(type_var_sp[a]))
        b +=1
    a +=1
    
#print('-------------------total =--------',b , '--------------------------' )
log('-------------------> total = '+ str(b))
#print('\n\n')
log('\n')

#print('-----------------------------input variable-------------------------------')
log('-----------------------------pointer in input-------------------------------')
a = 3
b = 0
while a <= (number_colum_input - 1):
    if (not('Number' in type_var[a]) and ('[0' in type_var[a]) and (not 'AMIN_return' in type_var[a])):
        #print(type_var[a])
        log(str(type_var[a]))
        b +=1
    a +=1
    
#print('-------------------total =--------',b , '--------------------------' )
log('-------------------> total = '+ str(b))
#print('\n\n')
log('\n')

log('-----------------------------array in input-------------------------------')
a = 3
b = 0
while a <= (number_colum_input - 1):
    if (not('Number' in type_var[a]) and ('[' in type_var[a]) and (not 'AMIN_return' in type_var[a])):
        #print(type_var[a])
        log(str(type_var[a]))
        b +=1
    a +=1
    
#print('-------------------total =--------',b , '--------------------------' )
log('-------------------> total = '+ str(b))
#print('\n\n')
log('\n')
log('\n')


#--------------------------
#print('\n\n')
#print('\n\n')

#p = 0
#while p < line_of_test:
#    #print(p, '\t',string_copy[p])
#    p += 1
log('-------Finish--------')
log('\n')
now = datetime.datetime.now()
mins_t = now.strftime("%M")
sec_t = now.strftime("%S")

time_old = int(mins)*60 + int(sec)
time_new = int(mins_t)*60 + int(sec_t)

log(str(now) + '\n')
print('\n' + '\n' + "Total time Running = " + str(time_new - time_old) +  " seconds" + '\n')
log("Total time Running = " + str(time_new - time_old) +  " seconds" + '\n')
file_test_log.close()  
try:
    os.remove(des_excel_copy)
except:
    print('Cannot delete temp report file')
print('-------Finish--------')


#time.sleep(0.001)
