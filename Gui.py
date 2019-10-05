#!usr/bin/python
import codecs
import openpyxl
import sys, os
from time import gmtime, strftime
from pathlib import Path
from openpyxl.utils import get_column_letter, column_index_from_string
import array as arr
import numpy as np
from shutil import copyfile
from tkinter import *
from tkinter import messagebox
import random
import time
import subprocess
import os, shutil
from shutil import copyfile

import sys, string, os
#from new_colect_testcase import run
from copy_write_time_cover_log_csv_infor_init_stub import copy_file, run_macro
#copy_file(path_winAMS_t, name_function_t, source_function_t )
#---------------------------------------------------------------------


#--------------------------------------------------------------------
window = Tk()
 
window.title("Analyzis TestCase")
 
window.geometry('600x500')

lb_status = Label(window, text="Status", font=("Arial Bold", 15),fg="red")
lb_status.grid(column=0, row=0)

lbl = Label(window, text="Type path excel file report: ")
lbl.grid(column=0, row=2)
txt_path = Entry(window,width=100)
txt_path.grid(column=0, row=3)

lb2 = Label(window, text="Type path Source .c test: ")
lb2.grid(column=0, row=5)
txt_func = Entry(window,width=70)
txt_func.grid(column=0, row=6)

path_excel = ''
source = ''
path_exe = r"D:/ver1.2.exe"
path = ''
#

def add_last_character(path):
    wb = openpyxl.load_workbook(path)
    sheet = wb['テストケース表']
    max_rown = sheet.max_row
    sheet['B' + str(max_rown +1)] = 'end'
    wb.save(path)

def task():
    box_1 = txt_path.get()
    box_1 = box_1.replace('\\','/')
    box_1 = box_1.replace('"','')
    box_2 = txt_func.get()
    lb_status.configure(text="Running Colect TestCase")
    # call exe colect testcase
    shutil.copy(box_1, 'D:/' )
    tmp = box_1.split('/')
    box_1 = 'D:/' + tmp[-1]
    add_last_character(box_1)
    subprocess.call([path_exe, box_1, box_2])
    os.remove(box_1)
    #run(box_1, box_2, 0, 0)
    lb_status.configure(text="Finish")
    txt_path.delete(0, END)
    txt_func.delete(0, END)
    
def clicked():
    task()

btn = Button(window, text="Colect testcase...", command=clicked,bg="orange", fg="red")
btn.grid(column=0, row=20)

lb_status = Label(window, text="Copy file and run Macro", font=("Arial Bold", 15),fg="blue")
lb_status.grid(column=0, row=21)

lb3 = Label(window, text="WinAms: ")
lb3.grid(column=0, row=22)
txt_WinAms = Entry(window,width=70)
txt_WinAms.grid(column=0, row=23)

lb4 = Label(window, text="Source.c: ")
lb4.grid(column=0, row=24)
txt_Source = Entry(window,width=70)
txt_Source.grid(column=0, row=25)

lb5 = Label(window, text="Function: ")
lb5.grid(column=0, row=26)
txt_Function = Entry(window,width=70)
txt_Function.grid(column=0, row=27)

return_path = ''
path = ''
path_t = ''

def clicked_copy():
    box_1 = txt_WinAms.get()
    box_1 = box_1.replace('\\','/')
    box_2 = txt_Source.get()
    box_3 = txt_Function.get()
    try:
        return_path = copy_file(box_1, box_3, box_2)
    except:
        print('can not copy file\n')
    path = return_path.split('$')
    print(path)
                                
def clicked_run():
    try:
        run_macro()
    except:
        print('can not run Macro\n')

temp = ''   
def clicked_colect():
    w = openpyxl.load_workbook("D:/temp.xlsx")
    sheet7 = w['Sheet1']
    temp = str(sheet7['A15'].value)
    temp = temp.replace('\\','/')
    temp = temp.replace('"','')
    temp = temp.split('$')
    box_1 = temp[0]
    box_2 = temp[1]
    #box_2 = box_2.replace(' ','')
    #box_1 = path[0]
    #box_2 = path[1]
    print(box_1)
    print(box_2)
    txt_path.delete(0, END)
    txt_path.insert(0,box_1)
    txt_func.delete(0, END)
    txt_func.insert(0,box_2)
    #subprocess.call([path_exe, box_1, box_2])
    
#------------------------------------------------------
b6 = Label(window, text="                 ")
b6.grid(column=0, row=28)
btn_1 = Button(window, text="Copy file...", command=clicked_copy,bg="orange", fg="green")
btn_1.grid(column=0, row=29)

b7 = Label(window, text="                 ")
b7.grid(column=0, row=30)
btn_2 = Button(window, text="run Macro...", command=clicked_run,bg="orange", fg="blue")
btn_2.grid(column=0, row=31)

b8 = Label(window, text="                 ")
b8.grid(column=0, row=32)
btn_3 = Button(window, text="Give path Tescase...", command=clicked_colect,bg="orange", fg="brown")
btn_3.grid(column=0, row=33)


#while 1:
window.mainloop()
 #print('hello')
 #time.sleep(5)
