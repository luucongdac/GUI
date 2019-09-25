import os, shutil
from shutil import copyfile
from sys import exit
import codecs
import openpyxl
import sys, os
from time import gmtime, strftime
from pathlib import Path
import win32com.client

print('Copy 14 files to folder test, type 1 to run macro copy html or 0 to just copy file')
run_csv = input()




#------------------------------------------------------------------------------------------------------------
# WinAMS workspace
path_winAMS = "D://Workspace//P33A_QCV2_Vol2_Group7_Stub_20190822//P33A_QCV2_vol2_Group7//root//UnitTest//WinAMSTest"

# function name
name_function = 'vCIR_CalcNeutralYawRate'

# source name
source_function = 'cir_neutral.c'

#=================================
wb_read = openpyxl.load_workbook('C:/Users/daccl.hy/Desktop/1.xlsx')
sheet_read = wb_read['Name']
path_winAMS = sheet_read['B89'].value
path_winAMS = path_winAMS.replace("\\", '//')
source_function = sheet_read['B90'].value
name_function = sheet_read['B91'].value
print(path_winAMS, source_function,  name_function)
#------------------------------



#workspace ouput release
path_out = 'D://SVN HilCS//Deliverables//Branches//Task00129_P33A_QCV2'
path_out_report = ''
path_out_csv = ''

#Report + store csv
report = '単体テスト仕様書'
folder_report = '単体テスト結果'

#template report excel
template = 'D://template.xlsx'

# cai nay tam
basepath = 'C://Users//luuco//Desktop//raw'
name_function_report = name_function + '.xlsx'
name_function1 = name_function
source_function1 = source_function


#-------------------------------------------------------------------------------------------------------------
#basepath_des = path_raw + '//' + name_function


# check exis folder to copy csv and report\
# ko co thi se tao va cap nhat path_out
print('finding Task follow function and name')
print('.........')

check = 0;
with os.scandir(path_out) as entries1:
    for entry in entries1:
        if entry.is_dir():
            if (source_function + '_DacLuu') in entry.name:
                path_out = path_out + '//' + entry.name
                check = 1
                print(path_out)
                print('.........')

if not check:
    path_out = path_out + '//Task00xxx_GroupX_' + source_function + '_DacLuu'
    os.makedirs(path_out)
    os.makedirs(path_out + '//単体テスト仕様書')
    os.makedirs(path_out + '//単体テスト結果'  )
    os.makedirs(path_out + '//単体テスト結果' + '//' +  name_function )
    print('create Task folder: ' + path_out)
    print('.........')


                  
path_out_report = path_out + '//' + '単体テスト仕様書'
path_out_csv = path_out + '//' + '単体テスト結果' + '//' + name_function
print(path_out_csv)
print('.........')

#check excel co chua, chua co thi copy tu temp va doi ten
print('copy and rename excel')
print('.........')
    
check = 0;
with os.scandir(path_out_report) as entries2:
    for entry2 in entries2:
        if entry2.is_file():
            if name_function in entry2.name:
                print('Exist excel')
                print('.........')
                check =1;
if not check:
    shutil.copy(template, path_out_report)
    old_file = os.path.join(path_out_report, template)
    new_file = os.path.join(path_out_report, name_function + '.xlsx')
    os.rename(old_file, new_file)
    shutil.copy(path_out_report + '//' + 'template.xlsx', 'D://' )
    os.remove(path_out_report + '//' + 'template.xlsx')
    print('Done')
    print('.........')


#check folder csv ton tai chua de copy moi, co roi xoa de ghi de
print('ReWrite folder CSV')
print('.........')

if os.path.exists(path_out_csv):
    try:
        shutil.rmtree(path_out_csv)
        os.makedirs(path_out_csv)
    except:
        print('permission fail to delete folder')
else:
    os.makedirs(path_out_csv)
#-----------------------------------------------------------------------------------------
#copy all file to csv folder otput

#copy file stub
print('copy Stub file')
print('.........')
shutil.copy(path_winAMS + '//AMSTB_SrcFile.c', path_out_csv )

#copy 8 file csv
print('copy 8 files CSV & html')
print('.........')
with os.scandir(path_winAMS + '//' + 'TestCsv' ) as entries3:
    for entry3 in entries3:
        if entry3.is_file():
            #print(entry3.name)
            if name_function in entry3.name:
                shutil.copy(entry3, path_out_csv )

#copy 2 file testReport & 2 file _Infor + _Table
print('copy 2 files _Info & _Table')
print('.........')
shutil.copy(path_winAMS + '//' + "Out2019-07-12(13'35'32)" + '//' + 'TestReport.csv', path_out_csv )
shutil.copy(path_winAMS + '//' + "Out2019-07-12(13'35'32)" + '//' + 'TestReport.htm', path_out_csv )
shutil.copy(path_winAMS + '//' + "Out2019-07-12(13'35'32)" + '//' + name_function + '_Info.html' , path_out_csv )
shutil.copy(path_winAMS + '//' + "Out2019-07-12(13'35'32)" + '//' + name_function + '_Table.html', path_out_csv )

#lay file coverlog.txt
print('copy Test coverlog')
print('.........')
if (len(name_function) > 20):
    cat = ''
    for x in range(20):
        cat = cat + name_function[x]
    name_function = cat

# copy test log to write to excel report
real_name_test_log = ''
with os.scandir(path_winAMS + '//' + "Out2019-07-12(13'35'32)" + '//' + 'TestCoverLog' + '//' + source_function ) as entries4:
    for entry4 in entries4:
        if entry4.is_file():
            #print(entry3.name)
            if name_function in entry4.name:
                shutil.copy(entry4, path_out_csv )
                real_name_test_log = entry4.name
                print(real_name_test_log)
                print(path_out_csv)
                print('.........')

#print(real_name_test_log)
print('copy all file done')
print('.........')
#print(path_out_report)


#--------------------write content to report file-----------------------------
#****************************************************************************
file = path_out_csv + '//' + real_name_test_log
excel = path_out_report + '//' + name_function_report
print(excel)
source_name = ''
function_name = ''
duong_link = ''
function_csv = ''

#------delete  end space \n \r -----
def delete(a):
    b = ''
    b = a.replace("\r",'')
    b = b.replace("\n",'')
    b = b.replace("\t",'')
    b = b.replace(' ','')
    b = b.replace('\"','')
    return(b)
def delete1(a):
    b = ''
    b = a.replace("\r",'')
    b = b.replace("\n",'')
    b = b.replace('\"','')
    return(b)

#---run csv
try: 
    xl=win32com.client.Dispatch('Excel.Application')
    xl.Workbooks.Open(Filename=r'D:\macro.xlsm', ReadOnly=1)
    #x_write.Workbooks.Open(Filename=r'C:\Users\daccl.hy\Desktop\1.xlsx', ReadOnly=1)
    if str(run_csv) == '1':
        xl.Application.Run("auto")
    #xl.Save()
    xl.Application.Quit()
    del xl
except:
    print('Fail run csv')

#-------
wb = openpyxl.load_workbook(excel)

#-----------------write date of testing----
#sheet = wb['フォーマット変更来歴']
#sheet['C5'] = strftime("%m/%d/%Y", gmtime())
#print('-------------------------')
#print('write date of testing Done')
#print('-------------------------')
#wb.save(excel)
#---------------write test cover log--------
#sheet = wb['カバレッジ結果']

temp1 = 'A'
temp2 = 5

chuoi_ten_function = ''
chuoi_ten_source = ''
C0 = ''
C1 = ''
MC_DC = ''
#for line in codecs.open(file, encoding="utf-8"):
for line in codecs.open(file, encoding="utf-8", errors = 'replace'):    #errors = 'replace'
    #print(line) 
    #sheet[temp1 + str(temp2)] = line
    if temp2 == 5:
        chuoi_ten_function = line
        #print(chuoi_ten_function)
    if temp2 == 6:
        chuoi_ten_source = line
        #print(chuoi_ten_source)
    if temp2 == 7:
        C0 = line
        #print(C0)
    if temp2 == 8:
        C1 = line
        #print(C1)
    if temp2 == 9:
        MC_DC = line
        #print(MC_DC)
    temp2 = temp2 + 1

#try:
#    for line in codecs.open(file, encoding="utf-8"):
#        j = 'check'
#except:
#    print('#####------------########')
#    print('-------fail font log ----')
#    print('-------------------------')
#    print('#####------------########')

print('write test cover log Done')
print('-------------------------')

#wb.save(excel)
#-----cat ten function & source to copy --------------
#Function name           : func_additional_signal_cal 
#Source file name        : D:\Workspace\P33A_QCV2_Vol1_Stub_Group6_20190719\P33A_QCV2_Group6_20190719\root\target\AD_Software_PLatform\swcCAM_RAD\workspace\CAM_RAD\RIR\add_processing.c 


x = chuoi_ten_function.split(': ')
#print(x)
x = x[1]
function_name = x
#print(function_name)  # function name
function_csv = function_name + '.csv'
function_csv = delete(function_csv)


x = chuoi_ten_source.replace("\\","/")
x = x.split('target')
x = x[1]
print(x)

y = ''
x = x.split('/')
count = 1
source_name = x[len(x) -1]
print(source_name)

while count < len(x) -1:
    if count == len(x) -2:
        y = y + x[count]
    else:
        y = y + x[count] + '/'
    count +=1
duong_link = y
print(y)

x = C0.split(': ')
C0 = delete(x[1])
x = C1.split(': ')
C1 = delete(x[1])
MC_DC = delete(MC_DC)
x = MC_DC.split(':')
MC_DC = x[1]
# check NG OK trong TestReport
path_report_csv = path_out_csv + '//' + 'TestReport.csv'
temp2 = 1
string = ''
for line in codecs.open(path_report_csv, encoding="utf-8", errors = 'replace'):    #errors = 'replace'
    if temp2 == 9:
        string = line.split(',')
        #print(string[1])
    temp2 = temp2 + 1

result_test = ''
bug_infor = ''
string[0] = delete(string[1])
if string[0] == 'Fault':
    result_test = 'NG'
    bug_infor = '29_P33A_QCV2_問題点管理表の問題点シートのNo'
else:
    result_test = 'OK'
    bug_infor = 'なし'
#print(result_test)
#print(string[0])
result_check = 'テスト結果: ' + result_test +'\n' + 'Ｃ０網羅率 : ' + C0 + '\n' + 'Ｃ１網羅率 : ' + C1 + '\n' + 'ＭＣ／ＤＣ網羅率 : ' + MC_DC + '\n' + delete1('問題点 : ' + bug_infor)
#---------------write infor test--------

sheet = wb['単体テスト仕様']

sheet['F8'] = delete(duong_link)
sheet['F9'] = delete(source_name)
#sheet['F10'] = delete(function_name)
sheet['F10'] = delete(name_function1)
sheet['F11'] = delete(name_function1) + '.csv' #function_csv
sheet['F12'] = result_check
wb.save(excel)

print('-------------------------')
print('write infor test Done')
print('-------------------------')
#--- check csv raw lay init & stub----------------------------------------------------------------
path_raw_csv = path_out_csv + '//' + name_function1 + '.csv'
temp2 = 1
string = ''
line_name_var_init = ''
line_value_var_init = ''

line_stub = ['']*30
line_stub_use = ['']*30
line_stub_none_use = ['']*30

temp2 = 0
temp1 = 1
line_name_var_init = ''
init_y_n = 0
for line in codecs.open(path_raw_csv, 'r' , encoding="utf-8", errors = 'replace'):    #errors = 'replace'
    #print(line)
    string = line.split(',')
    if string[0] == '#InitWheneverCall':
        line_name_var_init = line
        init_y_n = 1
        #print(temp1)
    if temp1 == 3:
        line_value_var_init = line
    if string[0] == '%':
        line_stub[temp2] = line
        temp2 = temp2 + 1   # number function stub
    temp1 += 1

if init_y_n == 1:    
    line_value_var_init_1 = line_value_var_init.split(',')
    #print(line_value_var_init_1)

    total_var_init = len(line_value_var_init_1)  #total var without first element

    line_name_var_init_1 = line_name_var_init.split(',')
else:
    total_var_init = 0
i=0
#print(total_var_init)
if total_var_init >0:
    while i< total_var_init:
        line_name_var_init_1[i] = line_name_var_init_1[i].replace("\"", '')
        i += 1
#print(line_name_var_init_1)

#---write init value to excel
if total_var_init > 0:   # neu co init
    print('------------------------------------Total init  = ',total_var_init -1)
    print('')
    #sheet = wb['単体テスト仕様']
    #sheet['B36'].value = ''

    k = 37
    #sheet.insert_rows(k,total_var_init - 1)   #insert bi merge
    i = 1
    while i < total_var_init:
        #sheet['G' + str(k)] =  delete(line_name_var_init_1[i])
        #print(delete(line_name_var_init_1[i]))
        #sheet['H' + str(k)] =  int(line_value_var_init_1[i])
        #print(int(line_value_var_init_1[i]))
        print(delete(line_name_var_init_1[i]) + '\t' + delete(line_value_var_init_1[i]))
        k +=1
        i +=1
    print('---------------------------------------------------------------------')
    print('')
print('-----------------------stub is use   ------------------------------------')
# in file stub
count = 0
for i in line_stub:
    if i != '':
        count += 1
#print(count)  #total stub

temp = ''
count1 = 0
if count >0:
    for i in line_stub:
        #print(i)
        temp = i.split(',')
        if temp[1] != '\"':
            #print('--- Stube is used---')
            print( delete1(temp[2] + '\t' + temp[1]))
        else:
            #print('--- Stube is NOT used---')
            print(delete1(temp[2] + '\t' + ' '))
        count1 +=1
        if count1 == count:
            break

print('------------------------------------------------------------------------------')
#-----------------save excel----------------

file = file.replace("//", "\\")
path_out_csv = path_out_csv.replace("//", "\\")
path_out_report = path_out_report.replace("//", "\\")
#print(file)
#print(path_out_csv)
wb.save(excel)
print('Save file.xlsx Done')
print('-------------------------')


#---------------------------------------------------
w = openpyxl.load_workbook("D:/temp.xlsx")

sheet7 = w['Sheet1']
sheet7['A1'] = name_function1
sheet7['A2'] = file
sheet7['A3'] = path_out_csv + '\\' + name_function1 + '_IE.html'
sheet7['A4'] = path_out_csv + '\\' + name_function1 + '_OE.html'
sheet7['A5'] = path_out_csv + '\\' + name_function1 + '_IO.html'
sheet7['A6'] = path_out_csv + '\\' + name_function1 + '_Table.html'

sheet7['A7'] = path_out_report + '\\' + name_function1 + '.xlsx'

sheet7['A8'] = real_name_test_log
sheet7['A9'] =  name_function1 + '_IE.html'
sheet7['A10'] = name_function1 + '_OE.html'
sheet7['A11'] =  name_function1 + '_IO.html'
sheet7['A12'] =  name_function1 + '_Table.html'
sheet7['A13'] = source_function1
path_winAMS = path_winAMS + 'AMSTB_SrcFile.c'
path_winAMS = path_winAMS.replace('//','\\')
sheet7['A14'] = path_winAMS
w.save("D:/temp.xlsx")

#---------------------------------------


input()
                
                

            
